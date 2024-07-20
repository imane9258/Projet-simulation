from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm 
from django.urls import reverse
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import Simulation

# Create your views here.


def connexion(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('accueil')
        else:
            messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
    return render(request, 'connexion.html')


from datetime import date, datetime
from decimal import Decimal
import json
from django.shortcuts import render
from django.utils.timezone import now
from django.db.models import Sum, Max
from .models import Simulation

def get_month_year(date):
    return date.strftime("%Y-%m")

def accueil(request):
    total_simulations = Simulation.objects.values('titre').distinct().count()
    total_investi = Simulation.objects.aggregate(Sum('prix_de_revient_total'))['prix_de_revient_total__sum']
    benefice = Simulation.objects.aggregate(Sum('marge_montant'))['marge_montant__sum']
    
    # Récupérer les dernières simulations par titre
    dernieres_simulations = Simulation.objects.values('titre').annotate(max_date=Max('date_creation')).order_by('-max_date')[:3]

    # Regrouper les simulations par titre
    simulations_grouped = {}
    for simulation in dernieres_simulations:
        titre = simulation['titre']
        details = Simulation.objects.filter(titre=titre).order_by('-titre')
        simulations_grouped[titre] = {'details': details}

    # Sélection de la période (mois ou année)
    selected_period = request.GET.get('period', 'month')
    current_year = now().year
    benefit_data = {}

    if selected_period == 'month':
        # Calcul des bénéfices mensuels
        simulations = Simulation.objects.filter(date_creation__year=current_year)
        for simulation in simulations:
            key = get_month_year(simulation.date_creation)
            if key not in benefit_data:
                benefit_data[key] = Decimal(0)
            benefit_data[key] += simulation.marge_montant

        # Préparer les données pour le graphique
        sorted_keys = sorted(benefit_data.keys())
        dates_benefice = [datetime.strptime(key, "%Y-%m").strftime("%Y-%m") for key in sorted_keys]
        data_benefice = [float(benefit_data[key]) for key in sorted_keys]

    elif selected_period == 'year':
        # Calcul des bénéfices annuels
        simulations = Simulation.objects.all()
        for simulation in simulations:
            key = simulation.date_creation.year
            if key not in benefit_data:
                benefit_data[key] = Decimal(0)
            benefit_data[key] += simulation.marge_montant

        # Préparer les données pour le graphique
        sorted_keys = sorted(benefit_data.keys())
        dates_benefice = [str(key) for key in sorted_keys]
        data_benefice = [float(benefit_data[key]) for key in sorted_keys]

    context = {
        'total_simulations': total_simulations,
        'total_investi': total_investi,
        'benefice': benefice,
        'simulations_grouped': simulations_grouped,
        'dates_benefice': json.dumps(dates_benefice),
        'data_benefice': json.dumps(data_benefice),
        'selected_period': selected_period
    }
    return render(request, 'accueil.html', context)


@login_required
def liste_simulations(request):
    # Retrieve all simulations
    simulations = Simulation.objects.all()

    # Group simulations by title
    grouped_simulations = {}
    for simulation in simulations:
        if simulation.titre not in grouped_simulations:
            grouped_simulations[simulation.titre] = {
                'id_simulation': simulation.id_simulation,
                'total_prix_avec_isb': 0,
                'lines': []
            }
        grouped_simulations[simulation.titre]['total_prix_avec_isb'] += simulation.prix_vente_total_ht_avec_isb
        grouped_simulations[simulation.titre]['lines'].append(simulation)

    context = {
        'grouped_simulations': grouped_simulations
    }

    return render(request, 'liste_simulations.html', context)


from decimal import Decimal
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from .models import Simulation

@login_required
@csrf_protect
def creer_simulation(request):
    if request.method == 'POST':
        titre = request.POST.get('titre')
        simulations_ids = []
        total_ht_devis = Decimal('0')
        total_ttc_devis = Decimal('0')
        marge_montant_total = Decimal('0')
        total_prix_revient = Decimal('0')
        total_isb = Decimal('0')
        total_tva = Decimal('0')

        data = request.POST.copy()
        del data['csrfmiddlewaretoken']

         

        for key, value in data.items():
            if key.startswith('designation') and value:
                index = key.replace('designation', '')
                try:
                    quantite = int(data[f'quantite{index}'])
                    prix_unitaire = Decimal(data[f'prix_unitaire{index}'])
                    frais_transit = Decimal(data[f'frais_transit{index}'])
                    frais_douane = Decimal(data[f'frais_douane{index}'])
                    marge_percentage = Decimal(data[f'marge_percentage{index}'])
                    pourcentage_banque = Decimal(data[f'pourcentage_banque{index}'])
                except (KeyError, ValueError):
                    return HttpResponse("Erreur de saisie : Veuillez vérifier que tous les champs sont remplis correctement.", status=400)

                # Calculs pour chaque ligne de simulation
                prix_total_achat = (quantite * prix_unitaire).quantize(Decimal('0.01'))
                pourcentage_banque_montant = ((prix_total_achat * pourcentage_banque) / Decimal('100.00')).quantize(Decimal('0.01'))
                prix_de_revient_total = (prix_total_achat + frais_transit + frais_douane + pourcentage_banque_montant).quantize(Decimal('0.01'))
                marge_montant = ((prix_de_revient_total * marge_percentage) / Decimal('100.00')).quantize(Decimal('0.01'))
                prix_vente_total_ht_sans_isb = (prix_de_revient_total + marge_montant).quantize(Decimal('0.01'))
    
                prix_vente_total_ht_avec_isb = ((prix_vente_total_ht_sans_isb) / Decimal('0.98')).quantize(Decimal('0.01'))
                isb_montant = (prix_vente_total_ht_avec_isb * Decimal('0.02')).quantize(Decimal('0.01'))
                prix_vente_total_ht = ((prix_vente_total_ht_avec_isb) / quantite).quantize(Decimal('0.01'))
              

                # Cumul des totaux pour toutes les lignes de simulation
                marge_montant_total += marge_montant
                total_prix_revient += prix_de_revient_total
                total_isb += isb_montant
                total_ht_devis += prix_vente_total_ht_avec_isb
                total_tva = (prix_vente_total_ht_avec_isb * Decimal('0.19')).quantize(Decimal('0.01'))
                total_ttc_devis = (total_ht_devis + total_tva).quantize(Decimal('0.01'))
  # Calcul du TTC et TVA basé sur le total HT
                
               
                # Création de l'instance Simulation
                simulation = Simulation(
                    titre=titre,
                    designation=value,
                    quantite=quantite,
                    prix=prix_unitaire,
                    prix_total_achat=prix_total_achat,
                    montant_transit=frais_transit,
                    montant_douane=frais_douane,
                    pourcentage_banque=pourcentage_banque,
                    pourcentage_banque_montant=pourcentage_banque_montant,
                    prix_de_revient_total=prix_de_revient_total,
                    marge_pourcentage=marge_percentage,
                    marge_montant=marge_montant,
                    isb=isb_montant,
                    prix_vente_total_ht_sans_isb=prix_vente_total_ht_sans_isb,
                    prix_vente_total_ht_avec_isb=prix_vente_total_ht_avec_isb,
                    prix_vente_total_ht=prix_vente_total_ht,
                    total_ht_devis=total_ht_devis,
                    total_ttc_devis=total_ttc_devis,
                    total_tva=total_tva,
                    marge_montant_total=marge_montant_total,
                    total_prix_revient=total_prix_revient,
                    total_isb=total_isb

                )
                simulation.save()
                simulations_ids.append(str(simulation.id_simulation))

        # Redirection vers la liste des simulations après sauvegarde
        ids_simulation = ','.join(simulations_ids)
        return redirect('resultat_simulation', ids_simulation=ids_simulation)

    return render(request, 'creer_simulation.html')

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from decimal import Decimal

@login_required
def resultat_simulation(request, ids_simulation):
    ids = ids_simulation.split(',')
    simulations = Simulation.objects.filter(id_simulation__in=ids)

    total_ht_devis = int(sum(simulation.prix_vente_total_ht_avec_isb for simulation in simulations))
    marge_montant_total = int(sum(simulation.marge_montant for simulation in simulations))
    total_prix_revient = int(sum(simulation.prix_de_revient_total for simulation in simulations))
    total_isb = int(sum(simulation.isb for simulation in simulations))

    # Calculating total_tva based on total_ht_devis
    total_tva = int(total_ht_devis * Decimal('0.19'))
    # Calculating total_ttc_devis based on total_ht_devis and total_tva
    total_ttc_devis = total_ht_devis + total_tva

    context = {
        'simulations': simulations,
        'total_ht_devis': total_ht_devis,
        'marge_montant_total': marge_montant_total,
        'total_prix_revient': total_prix_revient,
        'total_isb': total_isb,
        'total_tva': total_tva,
        'total_ttc_devis': total_ttc_devis,
        'simulation_titre': simulations[0].titre if simulations else 'Aucune simulation disponible',
        'ids_simulation': ids_simulation,  # Add this line to include ids_simulation in the context
    }
    return render(request, 'resultat_simulation.html', context)



from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.db.models import Sum
from decimal import Decimal
import io

@login_required
def download_pdf(request, ids_simulation):
    try:
        ids_simulations = [int(id_) for id_ in ids_simulation.split(',')]
        simulations = Simulation.objects.filter(id_simulation__in=ids_simulations)
    except ValueError:
        return HttpResponse("Invalid simulation IDs", status=400)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="detail_simulation_{ids_simulation}.pdf"'

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Title
    p.setFont("Helvetica-Bold", 18)
    title_text = "Résultats de la Simulation"
    p.drawCentredString(width / 2, height - 30, title_text)

    y_position = height - 60

    # Afficher le nom du client une seule fois
    if simulations.exists():
        p.setFont("Helvetica-Bold", 17)
        client_text = f"Client: {simulations.first().titre}"
        p.drawCentredString(width / 2, y_position, client_text)
        y_position -= 30

    # Create Sample Styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 12

    bold_style = ParagraphStyle(
        'Bold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
    )

    # Fonction pour formater les nombres
    def format_number(value):
        return f"{int(value):,}".replace(',', ' ')

    # Prepare list data
    for simulation in simulations:
        if y_position < 100:  # Check if there's enough space on the page
            p.showPage()  # Start a new page
            y_position = height - 50  # Reset the position

            # Draw the title again on the new page
            p.setFont("Helvetica-Bold", 18)
            p.drawCentredString(width / 2, height - 30, title_text)
            y_position = height - 60

            # Afficher le nom du client sur les nouvelles pages
            p.setFont("Helvetica-Bold", 17)
            p.drawCentredString(width / 2, y_position, client_text)
            y_position -= 30

        # Drawing data as list items
        data = [
            ("Désignation", simulation.designation),
            ("Prix d'Achat Unitaire", f"{format_number(simulation.prix)} FCFA"),
            ("Quantité", simulation.quantite),
            ("Prix Total Achat", f"{format_number(simulation.prix_total_achat)} FCFA"),
            ("Frais Transport", f"{format_number(simulation.montant_transit)} FCFA"),
            ("Frais Douane", f"{format_number(simulation.montant_douane)} FCFA"),
            ("Pourcentage Banque", f"{simulation.pourcentage_banque}%"),
            ("Montant Banque", f"{format_number(simulation.pourcentage_banque_montant)} FCFA"),
            ("Marge Pourcentage", f"{simulation.marge_pourcentage}%"),
            ("Montant Marge", f"{format_number(simulation.marge_montant)} FCFA"),
            ("Prix Vente Unitaire", f"{format_number(simulation.prix_vente_total_ht)} FCFA"),
            ("Prix de Revient", f"{format_number(simulation.prix_de_revient_total)} FCFA"),
            ("ISB", f"{format_number(simulation.isb)} FCFA"),
            ("Total HT avec ISB", f"{format_number(simulation.prix_vente_total_ht_avec_isb)} FCFA")
        ]

        for label, value in data:
            p.setFont("Helvetica-Bold", 10)
            p.drawString(20, y_position, f"{label} : ")
            p.setFont("Helvetica", 10)
            p.drawString(150, y_position, f"{value}")
            y_position -= 12  # Adjust line spacing as needed

        y_position -= 10  # Add space after each simulation

    # Totals Section
    if y_position < 100:  # Check if there's enough space for the totals section
        p.showPage()  # Start a new page
        y_position = height - 50  # Reset the table start position

        # Draw the title again on the new page
        p.setFont("Helvetica-Bold", 18)
        p.drawCentredString(width / 2, height - 30, title_text)
        y_position = height - 60

        # Afficher le nom du client sur les nouvelles pages
        p.setFont("Helvetica-Bold", 17)
        p.drawCentredString(width / 2, y_position, client_text)
        y_position -= 30

    # Draw the separator line before the totals
    p.setStrokeColorRGB(0, 0, 0)
    p.setLineWidth(1)
    p.line(20, y_position, width - 20, y_position)
    y_position -= 20

    p.setFont("Helvetica-Bold", 12)
    p.drawString(20, y_position, "Totaux")
    y_position -= 20

    # Calcul des totaux
    total_marge_montant = simulations.aggregate(Sum('marge_montant'))['marge_montant__sum'] or 0
    total_prix_revient = simulations.aggregate(Sum('prix_de_revient_total'))['prix_de_revient_total__sum'] or 0
    total_isb = simulations.aggregate(Sum('isb'))['isb__sum'] or 0
    total_ht_devis = simulations.aggregate(Sum('prix_vente_total_ht_avec_isb'))['prix_vente_total_ht_avec_isb__sum'] or 0
    total_tva = (total_ht_devis * Decimal('0.19')).quantize(Decimal('0.01'))
    total_ttc_devis = (total_ht_devis + total_tva).quantize(Decimal('0.01'))

    # Draw Totals
    totals = [
        ("Marge Montant Total", f"{format_number(total_marge_montant)} FCFA"),
        ("Total Prix Revient", f"{format_number(total_prix_revient)} FCFA"),
        ("ISB", f"{format_number(total_isb)} FCFA"),
        ("Total HT du Devis", f"{format_number(total_ht_devis)} FCFA"),
        ("TVA", f"{format_number(total_tva)} FCFA"),
        ("Total TTC du Devis", f"{format_number(total_ttc_devis)} FCFA")
    ]

    for label, value in totals:
        if y_position < 20:  # Check if there's enough space on the page
            p.showPage()  # Start a new page
            y_position = height - 50  # Reset the table start position

            # Draw the title again on the new page
            p.setFont("Helvetica-Bold", 18)
            p.drawCentredString(width / 2, height - 30, title_text)
            y_position = height - 60

            # Afficher le nom du client sur les nouvelles pages
            p.setFont("Helvetica-Bold", 17)
            p.drawCentredString(width / 2, y_position, client_text)
            y_position -= 30

        p.setFont("Helvetica-Bold", 10)
        p.drawString(20, y_position, f"{label} : ")
        p.setFont("Helvetica", 10)
        p.drawString(150, y_position, f"{value}")
        y_position -= 12  # Adjust line spacing as needed

    p.showPage()
    p.save()

    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()

    return response



import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.contrib.auth.decorators import login_required
from .models import Simulation
from decimal import Decimal
from django.db.models import Sum

@login_required
def download_excel(request, ids_simulation):
    try:
        simulation = Simulation.objects.get(id_simulation=ids_simulation)
    except Simulation.DoesNotExist:
        return HttpResponse("Simulation not found", status=404)

    # Fetch all simulation lines for the given title
    simulations_grouped_by_title = Simulation.objects.filter(titre=simulation.titre)

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Header 
    headers = [
        "Désignation", "Quantité", "Prix d'Achat Unitaire", "Prix Total Achat",
        "Frais Transit", "Frais Douane", "Pourcentage Banque", "Montant Banque",
        "Marge Pourcentage", "Montant Marge", "Prix Vente Unitaire", "Prix de Revient",
        "ISB", "Total HT avec ISB"
    ]
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 20

    # Function to replace dots with spaces
    def format_number(value):
        if value is not None:
            return str(int(value)).replace('.', ' ')
        return value

    # Add data
    row_num = 2
    for line in simulations_grouped_by_title:
        row = [
            line.designation,
            line.quantite,
            format_number(line.prix),
           
            format_number(line.prix_total_achat),
            format_number(line.montant_transit),
            format_number(line.montant_douane),
            format_number(line.pourcentage_banque) + "%",
            format_number(line.pourcentage_banque_montant),
            format_number(line.marge_pourcentage) + "%",
            format_number(line.marge_montant),
            format_number(line.prix_vente_total_ht),
            format_number(line.prix_de_revient_total),
            format_number(line.isb),
            format_number(line.prix_vente_total_ht_avec_isb)
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        row_num += 1

    # Totals
    total_font = Font(size=12, bold=True)
    total_fill = PatternFill(start_color="B0F2B6", end_color="B0F2B6", fill_type="solid")
    # Calcul des totaux pour les simulations avec le même titre
    marge_montant_total = simulations_grouped_by_title.aggregate(Sum('marge_montant'))['marge_montant__sum']
    total_prix_revient = simulations_grouped_by_title.aggregate(Sum('prix_de_revient_total'))['prix_de_revient_total__sum']
    total_isb = simulations_grouped_by_title.aggregate(Sum('isb'))['isb__sum']
    total_ht_devis = simulations_grouped_by_title.aggregate(Sum('prix_vente_total_ht_avec_isb'))['prix_vente_total_ht_avec_isb__sum']
    
    # Calculate total_tva and total_ttc_devis
    total_tva = (total_ht_devis * Decimal('0.19')).quantize(Decimal('0.01'))
    total_ttc_devis = (total_ht_devis + total_tva).quantize(Decimal('0.01'))

    # Format totals to string with spaces
    formatted_marge_montant_total = format_number(marge_montant_total)
    formatted_total_prix_revient = format_number(total_prix_revient)
    formatted_total_isb = format_number(total_isb)
    formatted_total_ht_devis = format_number(total_ht_devis)
    formatted_total_tva = format_number(total_tva)
    formatted_total_ttc_devis = format_number(total_ttc_devis)

    # Add totals to the bottom of respective columns
    ws.cell(row=row_num, column=10, value=formatted_marge_montant_total).font = total_font  # Marge Montant Total
    ws.cell(row=row_num, column=10).fill = total_fill
    ws.cell(row=row_num, column=12, value=formatted_total_prix_revient).font = total_font  # Total Prix Revient
    ws.cell(row=row_num, column=12).fill = total_fill
    ws.cell(row=row_num, column=13, value=formatted_total_isb).font = total_font  # ISB
    ws.cell(row=row_num, column=13).fill = total_fill
    
    # Totals with labels
    ws.cell(row=row_num, column=14, value="Total HT Devis").font = total_font
    ws.cell(row=row_num, column=14).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num, column=14).fill = total_fill
    ws.cell(row=row_num, column=15, value=formatted_total_ht_devis).font = total_font
    ws.cell(row=row_num, column=15).fill = total_fill

    ws.cell(row=row_num + 1, column=14, value="TVA").font = total_font
    ws.cell(row=row_num + 1, column=14).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num + 1, column=14).fill = total_fill
    ws.cell(row=row_num + 1, column=15, value=formatted_total_tva).font = total_font
    ws.cell(row=row_num + 1, column=15).fill = total_fill

    ws.cell(row=row_num + 2, column=14, value="Total TTC Devis").font = total_font
    ws.cell(row=row_num + 2, column=14).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num + 2, column=14).fill = total_fill
    ws.cell(row=row_num + 2, column=15, value=formatted_total_ttc_devis).font = total_font
    ws.cell(row=row_num + 2, column=15).fill = total_fill

    # Save the workbook to a BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the response as an Excel file
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="detail_simulation_{ids_simulation}.xlsx"'

    return response


# views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Simulation
from .forms import SimulationForm
from .utils import recalculer_simulation



# views.py

from django.shortcuts import render, get_object_or_404
from django.db.models import Sum
from .models import Simulation

def detail_simulation(request, id_simulation):
    simulation = get_object_or_404(Simulation, id_simulation=id_simulation)

    simulations_grouped_by_title = Simulation.objects.filter(titre=simulation.titre)
    
   

    # Calcul des totaux pour les simulations avec le même titre
    marge_montant_total = simulations_grouped_by_title.aggregate(Sum('marge_montant'))['marge_montant__sum']
    total_prix_revient = simulations_grouped_by_title.aggregate(Sum('prix_de_revient_total'))['prix_de_revient_total__sum']
    total_isb = simulations_grouped_by_title.aggregate(Sum('isb'))['isb__sum']
    total_ht_devis = simulations_grouped_by_title.aggregate(Sum('prix_vente_total_ht_avec_isb'))['prix_vente_total_ht_avec_isb__sum']
    total_tva = (total_ht_devis * Decimal('0.19')).quantize(Decimal('0.01'))
    total_ttc_devis = (total_ht_devis + total_tva).quantize(Decimal('0.01'))

    context = {
        'simulation': simulation,
        'titre': simulation.titre,
        'simulations': simulations_grouped_by_title,
        'id_simulation': id_simulation,
        'marge_montant_total': marge_montant_total,
        'total_prix_revient': total_prix_revient,
        'total_isb': total_isb,
        'total_ht_devis': total_ht_devis,
        'total_tva': total_tva,
        'total_ttc_devis': total_ttc_devis,
       
    }
    return render(request, 'detail_simulation.html', context)


from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from .models import Simulation

def edit_simulation(request, id_simulation):
    simulation = get_object_or_404(Simulation, id_simulation=id_simulation)
    simulations_grouped_by_title = Simulation.objects.filter(titre=simulation.titre)

    if request.method == 'POST':
        titre = request.POST.get('titre')

        try:
            for line in simulations_grouped_by_title:
                designation = request.POST.get(f'designation_{line.id_simulation}')
                quantite = request.POST.get(f'quantite_{line.id_simulation}')
                prix_unitaire = request.POST.get(f'prix_unitaire_{line.id_simulation}')
                frais_transit = request.POST.get(f'frais_transit_{line.id_simulation}')
                frais_douane = request.POST.get(f'frais_douane_{line.id_simulation}')
                marge_percentage = request.POST.get(f'marge_percentage_{line.id_simulation}')
                pourcentage_banque = request.POST.get(f'pourcentage_banque_{line.id_simulation}')

                # Remplacer les virgules par des points pour les valeurs numériques
                quantite = quantite.replace(',', '.')
                prix_unitaire = prix_unitaire.replace(',', '.')
                frais_transit = frais_transit.replace(',', '.')
                frais_douane = frais_douane.replace(',', '.')
                marge_percentage = marge_percentage.replace(',', '.')
                pourcentage_banque = pourcentage_banque.replace(',', '.')

                # Convertir les valeurs récupérées en Decimal
                try:
                    quantite = Decimal(quantite)
                    prix_unitaire = Decimal(prix_unitaire)
                    frais_transit = Decimal(frais_transit)
                    frais_douane = Decimal(frais_douane)
                    marge_percentage = Decimal(marge_percentage)
                    pourcentage_banque = Decimal(pourcentage_banque)
                except InvalidOperation:
                    return render(request, 'edit_simulation.html', {
                        'simulation': simulation,
                        'simulations_grouped_by_title': simulations_grouped_by_title,
                        'error': 'Invalid input. Please enter valid numeric values.'
                    })

                # Vérifier les valeurs pour éviter la division par zéro
                if quantite == 0 or prix_unitaire == 0:
                    return render(request, 'edit_simulation.html', {
                        'simulation': simulation,
                        'simulations_grouped_by_title': simulations_grouped_by_title,
                        'error': 'Quantité et Prix Unitaire doivent être supérieurs à zéro.'
                    })

                prix_total_achat = (quantite * prix_unitaire).quantize(Decimal('0.01'))
                pourcentage_banque_montant = ((prix_total_achat * pourcentage_banque) / Decimal('100.00')).quantize(Decimal('0.01'))
                prix_de_revient_total = (prix_total_achat + frais_transit + frais_douane + pourcentage_banque_montant).quantize(Decimal('0.01'))
                marge_montant = ((prix_de_revient_total * marge_percentage) / Decimal('100.00')).quantize(Decimal('0.01'))
                prix_vente_total_ht_sans_isb = (prix_de_revient_total + marge_montant).quantize(Decimal('0.01'))
                prix_vente_total_ht_avec_isb = (prix_vente_total_ht_sans_isb / Decimal('0.98')).quantize(Decimal('0.01'))
                isb_montant = (prix_vente_total_ht_avec_isb * Decimal('0.02')).quantize(Decimal('0.01'))
                prix_vente_total_ht = (prix_vente_total_ht_avec_isb / quantite).quantize(Decimal('0.01'))

                # Mettre à jour l'objet simulation avec les nouvelles valeurs
                line.titre = titre
                line.designation = designation
                line.quantite = quantite
                line.prix = prix_unitaire
                line.montant_transit = frais_transit
                line.montant_douane = frais_douane
                line.marge_pourcentage = marge_percentage
                line.pourcentage_banque = pourcentage_banque
                line.prix_total_achat = prix_total_achat
                line.pourcentage_banque_montant = pourcentage_banque_montant
                line.prix_de_revient_total = prix_de_revient_total
                line.marge_montant = marge_montant
                line.prix_vente_total_ht_sans_isb = prix_vente_total_ht_sans_isb
                line.prix_vente_total_ht_avec_isb = prix_vente_total_ht_avec_isb
                line.isb_montant = isb_montant
                line.prix_vente_total_ht = prix_vente_total_ht

                line.save()

        except InvalidOperation as e:
            return render(request, 'edit_simulation.html', {
                'simulation': simulation,
                'simulations_grouped_by_title': simulations_grouped_by_title,
                'error': f'An error occurred during calculations: {e}'
            })

        return redirect('detail_simulation', id_simulation=simulation.id_simulation)

    return render(request, 'edit_simulation.html', {'simulation': simulation, 'simulations_grouped_by_title': simulations_grouped_by_title})
from django.shortcuts import render
from django.db.models import Sum
from .models import Simulation
from datetime import timedelta
from django.utils.timezone import now
from decimal import Decimal

def rapport_simulation(request):
    periode = request.GET.get('periode')

    simulations = Simulation.objects.none()
    date_debut = None
    date_fin = None

    if periode == 'jour':
        date_debut = now().date()
        date_fin = date_debut + timedelta(days=1)
    elif periode == 'mois':
        date_debut = now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = date_debut.month + 1 if date_debut.month < 12 else 1
        next_year = date_debut.year if next_month != 1 else date_debut.year + 1
        date_fin = date_debut.replace(month=next_month, year=next_year)
    elif periode == 'annee':
        date_debut = now().replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_fin = date_debut.replace(year=date_debut.year + 1)

    if date_debut and date_fin:
        simulations = Simulation.objects.filter(date_creation__range=[date_debut, date_fin])

    grouped_simulations = {}
    for simulation in simulations:
        key = (simulation.titre, simulation.date_creation.strftime("%Y-%m-%d %H:%M"))
        if key not in grouped_simulations:
            grouped_simulations[key] = {
                'titre': simulation.titre,
                'date_creation': simulation.date_creation,
                'total_ttc': 0,
                'total_marge': 0,
                'simulations': []
            }
        grouped_simulations[key]['total_ttc'] += simulation.prix_vente_total_ht_avec_isb * Decimal('1.19')
        grouped_simulations[key]['total_marge'] += simulation.marge_montant
        grouped_simulations[key]['simulations'].append(simulation)

    context = {
        'grouped_simulations': grouped_simulations.values(),
    }

    return render(request, 'rapport_simulation.html', context)

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from .models import Simulation
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from django.db.models import Sum
import io

def format_number(value):
    """Format number with space for thousands and no decimal places."""
    if value is None:
        value = Decimal(0)
    formatted_number = f"{int(value):,}".replace(",", " ")
    return formatted_number + " FCFA"

@login_required
def download_detail_pdf(request, id_simulation):
    try:
        simulation = Simulation.objects.get(id_simulation=id_simulation)
    except Simulation.DoesNotExist:
        return HttpResponse("Simulation not found", status=404)

    # Fetch all simulation lines for the given title
    simulations_grouped_by_title = Simulation.objects.filter(titre=simulation.titre)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="detail_simulation_{id_simulation}.pdf"'

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)

    width, height = A4

    # Title
    p.setFont("Helvetica-Bold", 18)
    title_text = f"Détails de la Simulation {id_simulation}"
    p.drawCentredString(width / 2, height - 30 * mm, title_text)

    # Client Title
    p.setFont("Helvetica", 12)
    client_title_text = f"Client: {simulation.titre}"
    p.drawString(10 * mm, height - 40 * mm, client_title_text)

    y_position = height - 50 * mm

    # Create styles
    styles = getSampleStyleSheet()
    normal_style = styles['BodyText']
    bold_style = ParagraphStyle(name='Bold', parent=normal_style, fontName='Helvetica-Bold', spaceAfter=6)

    for line in simulations_grouped_by_title:
        data = [
            ("Désignation", line.designation),
            ("Prix d'Achat Unitaire", format_number(line.prix)),
            ("Quantité", line.quantite),
            ("Prix Total Achat", format_number(line.prix_total_achat)),
            ("Frais Transit", format_number(line.montant_transit)),
            ("Frais Douane", format_number(line.montant_douane)),
            ("Pourcentage Banque", f"{line.pourcentage_banque}%"),
            ("Montant Banque", format_number(line.pourcentage_banque_montant)),
            ("Marge Pourcentage", f"{line.marge_pourcentage}%"),
            ("Montant Marge", format_number(line.marge_montant)),
            ("Prix Vente Unitaire", format_number(line.prix_vente_total_ht)),
            ("Prix de Revient", format_number(line.prix_de_revient_total)),
            ("ISB", format_number(line.isb)),
            ("Total HT avec ISB", format_number(line.prix_vente_total_ht_avec_isb))
        ]

        for item in data:
            title_paragraph = Paragraph(f"{item[0]}:", bold_style)
            value_paragraph = Paragraph(f"{item[1]}", normal_style)

            w, h = title_paragraph.wrap(width - 20 * mm, y_position)
            title_paragraph.drawOn(p, 10 * mm, y_position - h)
            w, h = value_paragraph.wrap(width - 40 * mm, y_position)
            value_paragraph.drawOn(p, 50 * mm, y_position - h)
            y_position -= h + 2 * mm

            if y_position < 20 * mm:  # Check if there's enough space on the page
                p.showPage()  # Start a new page
                y_position = height - 20 * mm  # Reset the table start position

                # Draw the title again on the new page
                p.setFont("Helvetica-Bold", 18)
                p.drawCentredString(width / 2, height - 30 * mm, title_text)
                p.setFont("Helvetica", 12)
                p.drawString(10 * mm, height - 40 * mm, client_title_text)
                y_position -= 20 * mm

        y_position -= 10 * mm  # Add extra space between different simulations

    # Add a line before the Totals section
    p.setStrokeColorRGB(0, 0, 0)
    p.setLineWidth(1)
    p.line(10 * mm, y_position, width - 10 * mm, y_position)
    y_position -= 5 * mm

    # Totals Section
    p.setFont("Helvetica-Bold", 12)
    p.drawString(10 * mm, y_position, "Totaux")
    y_position -= 8 * mm

    # Calcul des totaux pour les simulations avec le même titre
    marge_montant_total = simulations_grouped_by_title.aggregate(Sum('marge_montant'))['marge_montant__sum'] or Decimal(0)
    total_prix_revient = simulations_grouped_by_title.aggregate(Sum('prix_de_revient_total'))['prix_de_revient_total__sum'] or Decimal(0)
    total_isb = simulations_grouped_by_title.aggregate(Sum('isb'))['isb__sum'] or Decimal(0)
    total_ht_devis = simulations_grouped_by_title.aggregate(Sum('prix_vente_total_ht_avec_isb'))['prix_vente_total_ht_avec_isb__sum'] or Decimal(0)
    total_tva = total_ht_devis * Decimal('0.19')
    total_ttc_devis = total_ht_devis + total_tva

    # Convert totals to integers to remove decimals
    total_tva = int(total_tva)
    total_ttc_devis = int(total_ttc_devis)

    # Draw Totals
    totals = [
        ("Marge Montant Total", format_number(marge_montant_total)),
        ("Total Prix Revient", format_number(total_prix_revient)),
        ("ISB", format_number(total_isb)),
        ("Total HT du Devis", format_number(total_ht_devis)),
        ("TVA", format_number(total_tva)),
        ("Total TTC du Devis", format_number(total_ttc_devis))
    ]

    for item in totals:
        title_paragraph = Paragraph(f"{item[0]}:", bold_style)
        value_paragraph = Paragraph(f"{item[1]}", normal_style)

        w, h = title_paragraph.wrap(width - 20 * mm, y_position)
        title_paragraph.drawOn(p, 10 * mm, y_position - h)
        w, h = value_paragraph.wrap(width - 40 * mm, y_position)
        value_paragraph.drawOn(p, 50 * mm, y_position - h)
        y_position -= h + 2 * mm

        if y_position < 20 * mm:  # Check if there's enough space on the page
            p.showPage()  # Start a new page
            y_position = height - 20 * mm  # Reset the table start position

            # Draw the title again on the new page
            p.setFont("Helvetica-Bold", 18)
            p.drawCentredString(width / 2, height - 30 * mm, title_text)
            p.setFont("Helvetica", 12)
            p.drawString(10 * mm, height - 40 * mm, client_title_text)
            y_position -= 20 * mm

    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')




import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.contrib.auth.decorators import login_required
from .models import Simulation
from decimal import Decimal
from django.db.models import Sum

@login_required
def download_detail_excel(request, id_simulation):
    try:
        simulation = Simulation.objects.get(id_simulation=id_simulation)
    except Simulation.DoesNotExist:
        return HttpResponse("Simulation not found", status=404)

    # Fetch all simulation lines for the given title
    simulations_grouped_by_title = Simulation.objects.filter(titre=simulation.titre)

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Header 
    headers = [
        "Désignation", "Quantité", "Prix d'Achat Unitaire", "Prix Total Achat",
        "Frais Transit", "Frais Douane", "Pourcentage Banque", "Montant Banque",
        "Marge Pourcentage", "Montant Marge", "Prix Vente Unitaire", "Prix de Revient",
        "ISB", "Total HT avec ISB"
    ]
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 20

    # Function to replace dots with spaces
    def format_number(value):
        if value is not None:
            return str(int(value)).replace('.', ' ')
        return value

    # Add data
    row_num = 2
    for line in simulations_grouped_by_title:
        row = [
            line.designation,
            line.quantite,
            format_number(line.prix),
           
            format_number(line.prix_total_achat),
            format_number(line.montant_transit),
            format_number(line.montant_douane),
            format_number(line.pourcentage_banque) + "%",
            format_number(line.pourcentage_banque_montant),
            format_number(line.marge_pourcentage) + "%",
            format_number(line.marge_montant),
            format_number(line.prix_vente_total_ht),
            format_number(line.prix_de_revient_total),
            format_number(line.isb),
            format_number(line.prix_vente_total_ht_avec_isb)
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        row_num += 1

    # Totals
    total_font = Font(size=12, bold=True)
    total_fill = PatternFill(start_color="B0F2B6", end_color="B0F2B6", fill_type="solid")
    # Calcul des totaux pour les simulations avec le même titre
    marge_montant_total = simulations_grouped_by_title.aggregate(Sum('marge_montant'))['marge_montant__sum']
    total_prix_revient = simulations_grouped_by_title.aggregate(Sum('prix_de_revient_total'))['prix_de_revient_total__sum']
    total_isb = simulations_grouped_by_title.aggregate(Sum('isb'))['isb__sum']
    total_ht_devis = simulations_grouped_by_title.aggregate(Sum('prix_vente_total_ht_avec_isb'))['prix_vente_total_ht_avec_isb__sum']
    
    # Calculate total_tva and total_ttc_devis
    total_tva = (total_ht_devis * Decimal('0.19')).quantize(Decimal('0.01'))
    total_ttc_devis = (total_ht_devis + total_tva).quantize(Decimal('0.01'))

    # Format totals to string with spaces
    formatted_marge_montant_total = format_number(marge_montant_total)
    formatted_total_prix_revient = format_number(total_prix_revient)
    formatted_total_isb = format_number(total_isb)
    formatted_total_ht_devis = format_number(total_ht_devis)
    formatted_total_tva = format_number(total_tva)
    formatted_total_ttc_devis = format_number(total_ttc_devis)

    # Add totals to the bottom of respective columns
    ws.cell(row=row_num, column=10, value=formatted_marge_montant_total).font = total_font  # Marge Montant Total
    ws.cell(row=row_num, column=10).fill = total_fill
    ws.cell(row=row_num, column=12, value=formatted_total_prix_revient).font = total_font  # Total Prix Revient
    ws.cell(row=row_num, column=12).fill = total_fill
    ws.cell(row=row_num, column=13, value=formatted_total_isb).font = total_font  # ISB
    ws.cell(row=row_num, column=13).fill = total_fill
    
    # Totals with labels
    ws.cell(row=row_num, column=14, value="Total HT Devis").font = total_font
    ws.cell(row=row_num, column=14).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num, column=14).fill = total_fill
    ws.cell(row=row_num, column=15, value=formatted_total_ht_devis).font = total_font
    ws.cell(row=row_num, column=15).fill = total_fill

    ws.cell(row=row_num + 1, column=14, value="TVA").font = total_font
    ws.cell(row=row_num + 1, column=14).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num + 1, column=14).fill = total_fill
    ws.cell(row=row_num + 1, column=15, value=formatted_total_tva).font = total_font
    ws.cell(row=row_num + 1, column=15).fill = total_fill

    ws.cell(row=row_num + 2, column=14, value="Total TTC Devis").font = total_font
    ws.cell(row=row_num + 2, column=14).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num + 2, column=14).fill = total_fill
    ws.cell(row=row_num + 2, column=15, value=formatted_total_ttc_devis).font = total_font
    ws.cell(row=row_num + 2, column=15).fill = total_fill

    # Save the workbook to a BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the response as an Excel file
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="detail_simulation_{id_simulation}.xlsx"'

    return response

